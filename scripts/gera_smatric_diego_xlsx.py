"""
Gera arquivos .xlsx (e .csv backup) para importação MANUAL no TOTVS RM:
- SMATRICPL_diego_2022_UN1.xlsx (1 linha)
- SMATRICULA_diego_2022_UN1.xlsx (13 linhas)

Dados extraídos das views export.smatricpl e export.smatricula filtrando Diego (RA 20142166) ano 2022.
Saída em data/exportacoes/2026-05-20/.

Uso: python scripts/gera_smatric_diego_xlsx.py
"""
import os
import subprocess
import csv
from pathlib import Path
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    raise SystemExit("Instale openpyxl: pip install openpyxl")


PROJ = Path(__file__).resolve().parent.parent
OUT = PROJ / "data" / "exportacoes" / "2026-05-20"
OUT.mkdir(parents=True, exist_ok=True)


def psql_query(query):
    """Executa query no Postgres e retorna lista de dicts. Strip \\r das células."""
    # Credenciais via CLAUDE.local.md (gitignored)
    local_md = (PROJ / "CLAUDE.local.md").read_text(encoding="utf-8")
    sec = local_md.split("## 1.")[1].split("## 2.")[0]
    pwd = [ln for ln in sec.split("\n") if "Password:" in ln][0].split("**")[2].strip()
    host = "192.168.1.91"
    user = "postgres"
    psql = r'C:\Program Files\PostgreSQL\18\bin\psql.exe'

    env = os.environ.copy()
    env["PGCLIENTENCODING"] = "LATIN1"
    env["PGPASSWORD"] = pwd

    cmd = [psql, "-h", host, "-U", user, "-d", "Edf_bd_legado",
           "-A", "-F", "|", "-c", query]
    out = subprocess.run(cmd, env=env, capture_output=True, text=True, encoding="latin-1")
    if out.returncode != 0:
        raise RuntimeError(f"psql falhou: {out.stderr}")
    lines = [l.replace("\r", "") for l in out.stdout.strip().split("\n") if l.strip()]
    if len(lines) < 2:
        return []
    # Header line + rows + footer "(N linhas)"
    header = lines[0].split("|")
    rows = []
    for line in lines[1:]:
        if line.startswith("(") and "linha" in line:
            break
        cells = line.split("|")
        if len(cells) == len(header):
            rows.append(dict(zip(header, cells)))
    return rows


def write_xlsx(out_path, sheet_name, header, rows, doc_rows=None):
    """Escreve arquivo .xlsx com cabeçalho destacado e linhas de dados."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Documentação no topo (se houver)
    start_row = 1
    if doc_rows:
        for txt in doc_rows:
            ws.cell(row=start_row, column=1, value=txt)
            ws.cell(row=start_row, column=1).font = Font(italic=True, color="666666")
            start_row += 1
        start_row += 1  # linha em branco

    # Header
    for ci, col in enumerate(header, 1):
        c = ws.cell(row=start_row, column=ci, value=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(horizontal="center")
    start_row += 1

    # Rows
    for row in rows:
        for ci, col in enumerate(header, 1):
            ws.cell(row=start_row, column=ci, value=row.get(col, ""))
        start_row += 1

    # Auto-fit (largura simples baseada em header)
    for ci, col in enumerate(header, 1):
        width = max(len(col) + 2, 12)
        max_value_len = max((len(str(r.get(col, ""))) for r in rows), default=0)
        width = max(width, min(max_value_len + 2, 40))
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = width

    wb.save(out_path)
    print(f"  OK {out_path.relative_to(PROJ)}")


def write_csv_backup(out_path, header, rows):
    """CSV LATIN1 com separador ';' (padrão TOTVS RM)."""
    with open(out_path, "w", encoding="latin-1", newline="", errors="replace") as f:
        w = csv.DictWriter(f, fieldnames=header, delimiter=";", extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in header})
    print(f"  OK {out_path.relative_to(PROJ)}")


# === SMATRICPL ===
print("=== SMATRICPL Diego 2022 ===")
smpl_rows = psql_query(
    "SELECT * FROM export.smatricpl "
    "WHERE \"RA\"='20142166' AND \"CODPERLET\"='2022'"
)
print(f"  Rows na view: {len(smpl_rows)}")

# Layout TOTVS RM oficial pra SMATRICPL — campos críticos primeiro
SMPL_HEADER = [
    "CODCOLIGADA", "IDPERLET", "IDHABILITACAOFILIAL", "RA",
    "CODFILIAL", "CODTIPOCURSO", "CODCURSO", "CODHABILITACAO", "CODGRADE",
    "TURNO", "CODTURMA", "CODPERLET",
    "STATUS", "STATUSRES",
    "DTMATRICULA", "DTRESULTADO",
    "PERIODO", "IDENTIFICADOR", "NUMCARTEIRA", "VIACARTEIRA",
    "NUMALUNO", "CARTEIRAEMITIDA", "DESCTIPOMAT",
]

# Diego 2022 valida: precisa preencher IDHABILITACAOFILIAL=24 (existe no RM) e
# garantir DTMATRICULA com hora
smpl_enriched = []
for r in smpl_rows:
    enriched = dict(r)
    enriched["IDHABILITACAOFILIAL"] = "24"  # IDHABFIL EF2-8-UN1-2022 já criado no RM
    enriched["IDPERLET"] = "12"             # IDPERLET 2022 Filial 1
    if enriched.get("DTMATRICULA") and "T" not in enriched["DTMATRICULA"]:
        enriched["DTMATRICULA"] = enriched["DTMATRICULA"] + " 00:00:00"
    smpl_enriched.append(enriched)

write_xlsx(
    OUT / "SMATRICPL_diego_2022_UN1.xlsx",
    "SMATRICPL",
    SMPL_HEADER,
    smpl_enriched,
    doc_rows=[
        "SMATRICPL - Matrícula no Período Letivo (1 row = 1 ano/aluno)",
        "Aluno: Diego Silva Pereira de Sousa | RA: 20142166 | Ano: 2022 | Turma: 8A EF2 | UN1",
        "Gerado em 2026-05-20 a partir de export.smatricpl",
        "Importar manualmente no TOTVS RM (EduMatricPLData via WS está bloqueado)",
        "IDHABILITACAOFILIAL=24 e IDPERLET=12 confirmados existentes no RM HOMOLOG",
    ]
)
write_csv_backup(
    OUT / "SMATRICPL_diego_2022_UN1.csv",
    SMPL_HEADER,
    smpl_enriched
)


# === SMATRICULA ===
print("\n=== SMATRICULA Diego 2022 (13 disciplinas) ===")
sma_rows = psql_query(
    "SELECT * FROM export.smatricula "
    "WHERE \"RA\"='20142166' AND \"CODPERLET\"='2022' "
    "ORDER BY \"CODDISC\"::int"
)
print(f"  Rows na view: {len(sma_rows)}")

# Mapping CODDISC -> IDTURMADISC (criados no RM nesta sessão)
# Ordem CODDISC -> IDTURMADISC capturada em 2026-05-19:
IDTURMADISC_MAP = {
    "7": "188",
    "19": "189",
    "21": "190",
    "32": "191",
    "49": "192",
    "51": "193",
    "55": "194",
    "62": "195",
    "67": "196",
    "76": "197",
    "84": "198",
    "85": "199",
    "104": "187",
}

SMA_HEADER = [
    "CODCOLIGADA", "IDPERLET", "IDTURMADISC", "IDHABILITACAOFILIAL", "RA",
    "CODFILIAL", "CODTIPOCURSO", "CODCURSO", "CODHABILITACAO", "CODGRADE",
    "TURNO", "CODTURMA", "CODPERLET", "CODDISC",
    "STATUS", "STATUSRES", "NUMDIARIO",
    "DTMATRICULA", "OBSHISTORICO", "TIPOMAT", "TIPODISCIPLINA",
    "DTALTERACAO", "DTALTERACAOSIST", "CODSUBTURMA",
    "NUMCREDITOSCOB", "COBPOSTERIORMATRIC",
    "CODTURMAORIGEM", "CODDISCORIGEM",
    "CODTURMAPRINCIPAL", "CODDISCPRINCIPAL",
]

sma_enriched = []
for r in sma_rows:
    enriched = dict(r)
    cod_disc = enriched.get("CODDISC", "")
    enriched["IDTURMADISC"] = IDTURMADISC_MAP.get(cod_disc, "")
    enriched["IDHABILITACAOFILIAL"] = "24"
    enriched["IDPERLET"] = "12"
    if enriched.get("DTMATRICULA") and " " in enriched["DTMATRICULA"] and "T" not in enriched["DTMATRICULA"]:
        enriched["DTMATRICULA"] = enriched["DTMATRICULA"].replace(" ", "T")
    sma_enriched.append(enriched)

write_xlsx(
    OUT / "SMATRICULA_diego_2022_UN1.xlsx",
    "SMATRICULA",
    SMA_HEADER,
    sma_enriched,
    doc_rows=[
        "SMATRICULA - Matrícula em Disciplina (13 rows = 13 disciplinas EF2 8º ano)",
        "Aluno: Diego Silva Pereira de Sousa | RA: 20142166 | Ano: 2022 | Turma: 8A EF2 | UN1",
        "Gerado em 2026-05-20 a partir de export.smatricula",
        "Importar APÓS SMATRICPL ser inserido manualmente",
        "IDTURMADISC mapeado para os IDs já criados no RM HOMOLOG (187-199)",
        "STATUS=Aprovado/STATUSRES=Aprovado (dados históricos 2022, ano já concluído)",
    ]
)
write_csv_backup(
    OUT / "SMATRICULA_diego_2022_UN1.csv",
    SMA_HEADER,
    sma_enriched
)


# === LEIA-ME ===
readme = OUT / "LEIA-ME_diego_2022.md"
readme.write_text(f"""# Importação Manual - Diego 2022 (EFII 8A UN1)

**Gerado em:** {date.today().isoformat()}
**Motivo:** `EduMatricPLData.SaveRecord` bloqueado via WS SOAP (resposta literal: "Você não está autorizado a inserir registros" em `ValidateInsertRecordSecurity()`). Reproduzido com user goliveira E com user consultor — não é perfil, é config global do DataServer. Aguardando intervenção do consultor.

## Arquivos

| Arquivo | Linhas | Importar em |
|---|---|---|
| `SMATRICPL_diego_2022_UN1.xlsx` | 1 | TOTVS RM → Educacional → Matrícula no Período Letivo |
| `SMATRICULA_diego_2022_UN1.xlsx` | 13 | TOTVS RM → Educacional → Matrícula em Disciplinas (APÓS SMATRICPL) |
| `*.csv` (LATIN-1, separador `;`) | igual | Backup pra import via SQL Server / outras ferramentas |

## Dados fonte

- View `export.smatricpl` (filtro RA='20142166' AND CODPERLET='2022')
- View `export.smatricula` (filtro RA='20142166' AND CODPERLET='2022')
- IDHABILITACAOFILIAL e IDPERLET enriquecidos com IDs já existentes no RM HOMOLOG

## Identificadores já existentes no RM (não criar de novo)

- **CODCOLIGADA:** 1
- **CODFILIAL:** 1 (UN1)
- **CODTIPOCURSO:** 1 (Educação Básica)
- **CODCURSO:** EF2
- **CODHABILITACAO:** 8 (8º ano)
- **CODGRADE:** 2022
- **CODTURMA:** 8A
- **CODPERLET:** 2022
- **IDPERLET:** 12 (Filial 1)
- **IDHABILITACAOFILIAL:** 24
- **IDTURMADISC:** 187-199 (mapeados por CODDISC na planilha)
- **RA aluno:** 20142166 (Diego)

## Ordem de importação obrigatória

1. **SMATRICPL** primeiro (1 row)
2. **SMATRICULA** depois (13 rows, pode ser em lote)

A inversão dessa ordem causa FK violation porque SMATRICULA referencia SMATRICPL (via RA + IDPERLET + IDHABILITACAOFILIAL).

## Verificação pós-importação

```sql
-- No banco RM Oracle (via SQL TOTVS)
SELECT COUNT(*) FROM SMATRICPL WHERE RA='20142166' AND CODPERLET='2022';
-- Esperado: 1

SELECT COUNT(*) FROM SMATRICULA WHERE RA='20142166' AND CODPERLET='2022';
-- Esperado: 13
```

Ou via WS SOAP (que tem leitura liberada):
```
ReadView EduMatricPLData filtro SMATRICPL.RA='20142166' AND SMATRICPL.CODPERLET='2022'
ReadView EduMatriculaData filtro SMATRICULA.RA='20142166' AND SMATRICULA.CODPERLET='2022'
```

## Próximo passo (após esses 2 importarem com sucesso)

Cadeia destravada:
- SCONTRATO × 4 (CODCONTRATO 2473, 2474, 2475, 2636) — via WS SaveRecord (já testado, funciona)
- SPARCELA × 37 — via WS SaveRecord
- SBOLSAALUNO × 4 (todas 100% filho de funcionário) — via WS SaveRecord
- SPROVAS/SNOTAS — após revisar view export.snotas inflada
""", encoding="utf-8")
print(f"\nOK {readme.relative_to(PROJ)}")

print("\nDONE.")
